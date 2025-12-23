import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PATH = "Voluntariado Base + WPForms - Areas (Dedup Nombre) - Pais Normalizado.xlsx"
SHEET = "Merged"
COL_AREAS = "Áreas de interés (lista)"
COL_PAIS = "País (normalizado)"
COL_NOMBRE = "Nombre completo"
COL_EMAIL = "Correo electrónico"
COL_TEL = "Teléfono"

# Mapeo de áreas TI
AREA_REDES = "Tecnología Sedes (Conectividad y Redes)"
AREA_DEPORTIVA = "Tecnología Deportiva"
AREA_SOPORTE = "Radios, Impresoras y Pantallas"

IT_AREAS = [AREA_REDES, AREA_DEPORTIVA, AREA_SOPORTE]

def load_data():
    try:
        df = pd.read_excel(PATH, sheet_name=SHEET)
    except Exception as e:
        print(f"Error cargando datos: {e}")
        return None
    
    # Normalizar nombre si no existe columna única
    if COL_NOMBRE not in df.columns and "Nombre completo: First" in df.columns:
        df[COL_NOMBRE] = df["Nombre completo: First"].fillna("") + " " + df["Nombre completo: Last"].fillna("")
    
    return df

def analyze_volunteer(row):
    areas_str = str(row.get(COL_AREAS, ""))
    user_areas = [x.strip() for x in areas_str.split(";") if x.strip()]
    
    # Flags booleanos
    is_redes = AREA_REDES in user_areas
    is_deportiva = AREA_DEPORTIVA in user_areas
    is_soporte = AREA_SOPORTE in user_areas
    is_it = is_redes or is_deportiva or is_soporte
    
    # Origen
    pais = str(row.get(COL_PAIS, "")).strip().lower()
    is_local = pais in ["república dominicana", "republica dominicana", "dominican republic"]
    origen = "Local" if is_local else "Internacional"
    
    # Perfil Técnico
    # High: Tiene 2+ áreas de TI
    it_count = sum([is_redes, is_deportiva, is_soporte])
    # Exclusive: Solo tiene áreas TI y nada más
    is_exclusive = len(user_areas) == it_count and it_count > 0
    
    perfil = "General"
    if is_exclusive:
        perfil = "Técnico Puro"
    elif it_count >= 2:
        perfil = "Técnico Versátil"
        
    return pd.Series({
        "Is_IT": is_it,
        "Is_Redes": is_redes,
        "Is_Deportiva": is_deportiva,
        "Is_Soporte": is_soporte,
        "Origen": origen,
        "Perfil": perfil,
        "Total_Areas": len(user_areas)
    })

def format_excel(writer, df, sheet_name):
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]
    
    # Estilos
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Ajustar columnas
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(adjusted_width, 50)

def main():
    df = load_data()
    if df is None: return

    # Enriquecer dataset
    analysis = df.apply(analyze_volunteer, axis=1)
    df = pd.concat([df, analysis], axis=1)
    
    # Filtrar solo TI
    df_it = df[df["Is_IT"]].copy()
    
    # Seleccionar columnas finales
    base_cols = ["Origen", "Perfil", COL_NOMBRE, COL_EMAIL, COL_TEL, COL_PAIS, COL_AREAS]
    
    # Crear DataFrames específicos
    df_master = df_it[base_cols]
    df_redes = df_it[df_it["Is_Redes"]][base_cols]
    df_deportiva = df_it[df_it["Is_Deportiva"]][base_cols]
    df_soporte = df_it[df_it["Is_Soporte"]][base_cols]

    # Dashboard Data
    total_it = len(df_it)
    count_redes = len(df_redes)
    count_deportiva = len(df_deportiva)
    count_soporte = len(df_soporte)
    
    count_local = len(df_it[df_it["Origen"] == "Local"])
    count_intl = len(df_it[df_it["Origen"] == "Internacional"])
    
    dashboard_data = {
        "Métrica": [
            "Total Voluntarios TI", 
            "--- Desglose por Área ---",
            "Redes y Conectividad",
            "Tecnología Deportiva",
            "Soporte (Radios/Pantallas)",
            "--- Desglose por Origen ---",
            "Locales (RD)",
            "Internacionales"
        ],
        "Cantidad": [
            total_it,
            "",
            count_redes,
            count_deportiva,
            count_soporte,
            "",
            count_local,
            count_intl
        ]
    }
    df_dash = pd.DataFrame(dashboard_data)

    output_path = "Organizacion_Voluntarios_TI.xlsx"
    
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # 1. Dashboard
        format_excel(writer, df_dash, "Resumen Dashboard")
        
        # 2. Lista Maestra
        format_excel(writer, df_master, "Todos TI")
        
        # 3. Sub-listas
        format_excel(writer, df_redes, "Redes y Conectividad")
        format_excel(writer, df_deportiva, "Tecnologia Deportiva")
        format_excel(writer, df_soporte, "Soporte Usuario")

    print(f"Archivo generado exitosamente: {output_path}")
    print(f"Total TI: {total_it}")
    print(f"Locales: {count_local} | Internacionales: {count_intl}")

if __name__ == "__main__":
    main()
